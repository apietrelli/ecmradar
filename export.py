"""
ECM Export - Esporta il DB in CSV/Excel per analisi esterna
=============================================================
Genera file pronti per R, Python/pandas, Excel, Power BI.

Uso:
    # Export completo in CSV (una cartella con tutti i file)
    python export.py --format csv

    # Export in un unico Excel con fogli multipli
    python export.py --format xlsx

    # Export solo la vista flat denormalizzata
    python export.py --format csv --flat-only

    # Export filtrato per anno
    python export.py --format xlsx --year 2025
"""

import argparse
import csv
import sqlite3
import logging
from pathlib import Path
from datetime import datetime

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("ecm_export")


# ── QUERY DI EXPORT ──────────────────────────────────────────────────────────

EXPORT_QUERIES = {
    "events": """
        SELECT e.*, p.name as provider_name
        FROM events e
        LEFT JOIN providers p ON e.provider_id = p.provider_id
        ORDER BY e.start_date DESC
    """,

    "providers": """
        SELECT p.*,
            COUNT(DISTINCT e.event_id) as total_events_actual,
            COUNT(DISTINCT s.sponsor_name) as n_sponsors,
            COUNT(DISTINCT sp.speaker_name) as n_speakers,
            GROUP_CONCAT(DISTINCT e.event_type) as event_types,
            GROUP_CONCAT(DISTINCT e.region) as regions
        FROM providers p
        LEFT JOIN events e ON p.provider_id = e.provider_id
        LEFT JOIN sponsors s ON e.event_id = s.event_id
        LEFT JOIN speakers sp ON e.event_id = sp.event_id
        GROUP BY p.provider_id
        ORDER BY total_events_actual DESC
    """,

    "sponsors": """
        SELECT s.sponsor_name,
            COUNT(DISTINCT s.event_id) as n_events,
            COUNT(DISTINCT e.provider_id) as n_providers,
            GROUP_CONCAT(DISTINCT e.event_type) as event_types,
            GROUP_CONCAT(DISTINCT e.region) as regions,
            ROUND(AVG(e.credits), 1) as avg_credits
        FROM sponsors s
        JOIN events e ON s.event_id = e.event_id
        GROUP BY s.sponsor_name
        ORDER BY n_events DESC
    """,

    "speakers": """
        SELECT sp.speaker_name, sp.role, sp.qualifica, sp.codice_fiscale,
            COUNT(DISTINCT sp.event_id) as n_events,
            COUNT(DISTINCT e.provider_id) as n_providers,
            GROUP_CONCAT(DISTINCT e.event_type) as event_types,
            GROUP_CONCAT(DISTINCT e.region) as regions
        FROM speakers sp
        JOIN events e ON sp.event_id = e.event_id
        GROUP BY sp.speaker_name
        ORDER BY n_events DESC
    """,

    "speaker_sponsor_links": """
        SELECT sp.speaker_name, sp.qualifica,
            s.sponsor_name,
            COUNT(DISTINCT e.event_id) as shared_events,
            GROUP_CONCAT(DISTINCT e.region) as regions,
            GROUP_CONCAT(DISTINCT e.objective) as objectives
        FROM speakers sp
        JOIN events e ON sp.event_id = e.event_id
        JOIN sponsors s ON e.event_id = s.event_id
        GROUP BY sp.speaker_name, s.sponsor_name
        ORDER BY shared_events DESC
    """,

    "sponsor_provider_matrix": """
        SELECT s.sponsor_name, p.name as provider_name,
            COUNT(DISTINCT e.event_id) as n_events,
            GROUP_CONCAT(DISTINCT e.event_type) as types,
            GROUP_CONCAT(DISTINCT e.region) as regions
        FROM sponsors s
        JOIN events e ON s.event_id = e.event_id
        JOIN providers p ON e.provider_id = p.provider_id
        GROUP BY s.sponsor_name, p.provider_id
        ORDER BY n_events DESC
    """,

    # Vista flat denormalizzata: una riga per ogni combinazione evento×speaker×sponsor
    "flat_view": """
        SELECT
            e.event_id,
            e.title,
            e.event_type,
            e.start_date,
            e.end_date,
            e.credits,
            e.cost,
            e.region,
            e.city,
            e.objective,
            e.max_participants,
            p.name as provider_name,
            p.provider_id,
            sp.speaker_name,
            sp.role as speaker_role,
            sp.qualifica as speaker_qualifica,
            sp.codice_fiscale as speaker_cf,
            s.sponsor_name
        FROM events e
        LEFT JOIN providers p ON e.provider_id = p.provider_id
        LEFT JOIN speakers sp ON e.event_id = sp.event_id
        LEFT JOIN sponsors s ON e.event_id = s.event_id
        ORDER BY e.start_date DESC, e.event_id, sp.speaker_name
    """
}


def export_csv(db_path: str, output_dir: str, year: int = 0, flat_only: bool = False):
    """Esporta in file CSV separati"""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    queries = {"flat_view": EXPORT_QUERIES["flat_view"]} if flat_only else EXPORT_QUERIES

    for name, query in queries.items():
        if year:
            # Aggiungi filtro anno se possibile
            if "e.start_date" in query:
                query = query.replace(
                    "ORDER BY",
                    f"HAVING 1=1 AND (e.start_date LIKE '%/{year}' OR e.start_date LIKE '{year}-%') ORDER BY"
                    if "GROUP BY" in query else
                    f"WHERE (e.start_date LIKE '%/{year}' OR e.start_date LIKE '{year}-%') ORDER BY"
                    if "WHERE" not in query else
                    f"AND (e.start_date LIKE '%/{year}' OR e.start_date LIKE '{year}-%') ORDER BY"
                )

        try:
            rows = conn.execute(query).fetchall()
        except Exception as e:
            log.warning(f"  Errore query {name}: {e}")
            # Riprova senza filtro anno
            rows = conn.execute(EXPORT_QUERIES[name]).fetchall()

        if not rows:
            log.info(f"  {name}: 0 righe, skip")
            continue

        filepath = out / f"{name}.csv"
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(rows[0].keys())
            for row in rows:
                writer.writerow(list(row))

        log.info(f"  {name}.csv → {len(rows)} righe")

    conn.close()
    log.info(f"\nCSV esportati in: {out}/")


def export_xlsx(db_path: str, output_path: str, year: int = 0, flat_only: bool = False):
    """Esporta in un unico file Excel con fogli multipli"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl non installato. Installa con: pip install openpyxl")
        log.info("Fallback: esporto in CSV...")
        export_csv(db_path, str(Path(output_path).parent / "ecm_export_csv"), year, flat_only)
        return

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    wb = Workbook()
    wb.remove(wb.active)  # rimuovi foglio vuoto

    # Stili
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="D9E2F3")
    )

    queries = {"flat_view": EXPORT_QUERIES["flat_view"]} if flat_only else EXPORT_QUERIES

    for name, query in queries.items():
        try:
            rows = conn.execute(query).fetchall()
        except Exception:
            rows = conn.execute(EXPORT_QUERIES.get(name, query)).fetchall()

        if not rows:
            continue

        # Nome foglio (max 31 char per Excel)
        sheet_name = name[:31].replace("_", " ").title()
        ws = wb.create_sheet(title=sheet_name)

        # Header
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Dati
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(list(row), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-width colonne
        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=1, column=col).value or "")),
                max((len(str(ws.cell(row=r, column=col).value or ""))
                     for r in range(2, min(len(rows) + 2, 100))),
                    default=10)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)

        # Filtri automatici
        ws.auto_filter.ref = ws.dimensions

        # Freeze prima riga
        ws.freeze_panes = "A2"

        log.info(f"  Foglio '{sheet_name}': {len(rows)} righe")

    conn.close()

    wb.save(output_path)
    log.info(f"\nExcel salvato: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="ECM Export")
    parser.add_argument("--db", default="ecm_database.db", help="Path database")
    parser.add_argument("--format", choices=["csv", "xlsx", "both"], default="both",
                        help="Formato di export")
    parser.add_argument("--output", default="", help="Path output (default: ecm_export_*)")
    parser.add_argument("--year", type=int, default=0, help="Filtra per anno")
    parser.add_argument("--flat-only", action="store_true",
                        help="Esporta solo la vista flat denormalizzata")

    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d")
    year_suffix = f"_{args.year}" if args.year else ""

    if args.format in ("csv", "both"):
        csv_dir = args.output or f"ecm_export_csv{year_suffix}_{timestamp}"
        log.info(f"Export CSV...")
        export_csv(args.db, csv_dir, args.year, args.flat_only)

    if args.format in ("xlsx", "both"):
        xlsx_path = args.output if args.output and args.output.endswith(".xlsx") \
            else f"ecm_export{year_suffix}_{timestamp}.xlsx"
        log.info(f"Export Excel...")
        export_xlsx(args.db, xlsx_path, args.year, args.flat_only)


if __name__ == "__main__":
    main()
